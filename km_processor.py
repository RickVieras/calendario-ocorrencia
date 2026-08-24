"""Processa a aba PROGRAMADO sem depender do Microsoft Excel."""
from __future__ import annotations
import io,re
from collections import defaultdict
from datetime import date,datetime
from pathlib import Path
import requests
from openpyxl import Workbook,load_workbook

COL_EMPRESA=3
COL_FROTA={"U":11,"S":12,"D":13}
COL_VIAGENS={"U":14,"S":15,"D":16}
COL_OPERACIONAL=17
COL_MORTA=18
COL_LINHA=4
COL_TRANSPORTA=22
CALENDARIO_INICIO=31
CALENDARIO_FIM=61
FIELDS=("frota","viagens","km_operacional","km_morta","km_transporta","km_total")

def number(value):
    if value is None or isinstance(value,bool): return 0.0
    try:return float(value)
    except(TypeError,ValueError):return 0.0

def to_date(value):
    if isinstance(value,datetime):return value.date()
    if isinstance(value,date):return value
    if isinstance(value,str):
        for fmt in("%Y-%m-%d","%d/%m/%Y","%d/%m/%y"):
            try:return datetime.strptime(value.strip(),fmt).date()
            except ValueError:pass
    return None

def is_transporta(company):
    return "transporta" in company.casefold()

def dates_in_period(ws,start_day,end_day):
    header=next(ws.iter_rows(min_row=2,max_row=2,values_only=True),())
    dates=[]
    for column in range(CALENDARIO_INICIO,CALENDARIO_FIM+1):
        current=to_date(header[column-1] if len(header)>=column else None)
        if current and start_day<=current.day<=end_day:dates.append((column,current))
    if not dates:raise ValueError("Não foram encontradas datas no calendário AE:BI.")
    return dates

def calculate(ws,start_day,end_day):
    dates=dates_in_period(ws,start_day,end_day)
    result=defaultdict(lambda:defaultdict(lambda:{field:0.0 for field in FIELDS}))
    types=defaultdict(lambda:defaultdict(set))
    for row in ws.iter_rows(min_row=3,values_only=True):
        raw=row[COL_EMPRESA-1] if len(row)>=COL_EMPRESA else None
        company=str(raw).strip() if raw is not None else ""
        if not company:continue
        linha=str(row[COL_LINHA-1] if len(row)>=COL_LINHA else "").strip()
        transporta=is_transporta(linha)
        if company.casefold() in {"none","nan","null","-","total","total geral","total por empresa"}:continue
        operational=number(row[COL_OPERACIONAL-1] if len(row)>=COL_OPERACIONAL else 0)
        dead_rate=number(row[COL_MORTA-1] if len(row)>=COL_MORTA else 0)
        for column,current in dates:
            schedule=str(row[column-1] if len(row)>=column else "").strip().upper()
            if schedule not in COL_FROTA:continue
            fleet=number(row[COL_FROTA[schedule]-1] if len(row)>=COL_FROTA[schedule] else 0)
            trips=number(row[COL_VIAGENS[schedule]-1] if len(row)>=COL_VIAGENS[schedule] else 0)
            op=number(row[COL_TRANSPORTA-1] if len(row)>=COL_TRANSPORTA else 0) if transporta else trips*operational
            km_transporta=0
            dead=fleet*dead_rate
            key=current.isoformat();entry=result[company][key]
            entry["frota"]+=fleet;entry["viagens"]+=trips;entry["km_operacional"]+=op
            entry["km_morta"]+=dead;entry["km_transporta"]+=km_transporta;entry["km_total"]+=op+dead+km_transporta
            types[company][key].add(schedule)
    if not result:raise ValueError("Nenhum dado de empresa foi encontrado na aba PROGRAMADO.")
    companies=[];all_daily=defaultdict(lambda:{field:0.0 for field in FIELDS});company_daily={}
    for company,rows in sorted(result.items()):
        total={"empresa":company,**{field:0.0 for field in FIELDS}};out=[]
        for key,values in sorted(rows.items()):
            out.append({"data":key,**values,"tipos":sorted(types[company][key])})
            for field in FIELDS:total[field]+=values[field];all_daily[key][field]+=values[field]
        companies.append(total);company_daily[company]=out
    daily=[{"data":key,**values}for key,values in sorted(all_daily.items())]
    return companies,daily,company_daily

def download_source(sheet_id):
    response=requests.get(f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx",timeout=90)
    response.raise_for_status();return response.content

def safe_name(name,used):
    base=re.sub(r'[\/:*?[]]'," ",name).strip()[:31] or "SEM EMPRESA";candidate=base;index=2
    while candidate.casefold() in used:
        suffix=f" ({index})";candidate=base[:31-len(suffix)]+suffix;index+=1
    used.add(candidate.casefold());return candidate

def write_daily_sheet(book,name,rows):
    ws=book.create_sheet(name)
    ws.append(["Data","Frota","Viagens","KM Operacional","KM Morta","KM Transporta","KM Total"])
    for item in rows:ws.append([item["data"],item["frota"],item["viagens"],item["km_operacional"],item["km_morta"],item["km_transporta"],item["km_total"]])
    ws.freeze_panes="A2";ws.auto_filter.ref=ws.dimensions
    for cell in ws[1]:cell.font=__import__("openpyxl").styles.Font(bold=True)
    for row in ws.iter_rows(min_row=2,min_col=1,max_col=1):
        row[0].number_format="dd/mm/yyyy"
    for column in "ABCDEFG":ws.column_dimensions[column].width=18
    return ws

def build_report_from_bytes(source,destination:Path,start_day=1,end_day=31):
    workbook=load_workbook(io.BytesIO(source),data_only=True,read_only=True)
    if "PROGRAMADO" not in workbook.sheetnames:raise ValueError("A planilha precisa ter a aba PROGRAMADO.")
    companies,daily,company_daily=calculate(workbook["PROGRAMADO"],start_day,end_day)
    output=Workbook();output.remove(output.active)
    summary=output.create_sheet("TOTAL POR EMPRESA")
    summary.append(["Empresa","Frota","Viagens","KM Operacional","KM Morta","KM Transporta","KM Total"])
    for item in companies:summary.append([item["empresa"],item["frota"],item["viagens"],item["km_operacional"],item["km_morta"],item["km_transporta"],item["km_total"]])
    write_daily_sheet(output,"TOTAL POR DIA",daily)
    used={"total por empresa","total por dia"}
    for company,rows in company_daily.items():write_daily_sheet(output,safe_name(company,used),rows)
    destination.parent.mkdir(parents=True,exist_ok=True);output.save(destination)
    return {"companies":companies,"daily":daily,"company_daily":company_daily}

def build_report(sheet_id,destination:Path,start_day=1,end_day=31):return build_report_from_bytes(download_source(sheet_id),destination,start_day,end_day)